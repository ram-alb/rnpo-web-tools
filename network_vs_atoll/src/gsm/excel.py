"""Fill excel file with gsm inconsistencies."""

from openpyxl import Workbook
from openpyxl.styles import Alignment


def fill_excel(gsm_diff, bsc):
    """
    Fill excel file with gsm inconsistencies.

    Args:
        gsm_diff: dict
        bsc: string

    Returns:
        string
    """
    columns = [
        'BSC',
        'Site',
        'Cell',
        'Parameter',
        'Network value',
        'Atoll value',
    ]

    work_book = Workbook()
    sheet = work_book.active

    for col in columns:
        current_cell = sheet.cell(row=1, column=columns.index(col) + 1)
        current_cell.value = col
        current_cell.alignment = Alignment(horizontal='center')

    bsc_diff = gsm_diff[bsc]
    row = 2
    for diff in bsc_diff:
        for column in columns:
            current_cell = sheet.cell(row=row, column=columns.index(column) + 1)
            current_cell.value = diff[column]
            current_cell.alignment = Alignment(horizontal='center')

        row += 1

    report_path = f'network_vs_atoll/reports/{bsc}-inconsistencies.xlsx'
    work_book.save(report_path)
    return report_path
