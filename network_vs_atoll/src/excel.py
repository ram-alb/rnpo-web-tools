"""Fill excel file with inconsistencies."""

from openpyxl import Workbook
from openpyxl.styles import Alignment


def get_colums_by_technology(technology):
    """
    Return columns name according to technology.

    Args:
        technology: string

    Returns:
        list
    """
    common_columns = [
        'Site',
        'Cell',
        'Parameter',
        'Network value',
        'Atoll value',
    ]

    columns_by_technology = {
        'LTE': ['Subnetwork', *common_columns],
        'WCDMA': ['RNC', *common_columns],
        'GSM': ['BSC', *common_columns],
    }

    return columns_by_technology[technology]


def fill_excel(technology, node_diffs, node):
    """
    Fill excel file with inconsistencies.

    Args:
        technology: string
        node_diffs: dict
        node: string

    Returns:
        string
    """
    work_book = Workbook()
    sheet = work_book.active

    columns = get_colums_by_technology(technology)
    for col in columns:
        current_cell = sheet.cell(row=1, column=columns.index(col) + 1)
        current_cell.value = col
        current_cell.alignment = Alignment(horizontal='center')

    row = 2
    for diff in node_diffs:
        for column in columns:
            current_cell = sheet.cell(row=row, column=columns.index(column) + 1)
            current_cell.value = diff[column]
            current_cell.alignment = Alignment(horizontal='center')
        row += 1

    report_path = f'network_vs_atoll/reports/{node}-inconsistencies.xlsx'
    work_book.save(report_path)
    return report_path
